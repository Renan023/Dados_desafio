import pandas as pd
import os
from openpyxl import load_workbook

def cadastro_excel(lista_pf, lista_pj, nome_arquivo='cadastro.xlsx'):
    # Verifica se o arquivo já existe
    arquivo_existe = os.path.exists(nome_arquivo)

    # Inicia o writer com modo correto
    with pd.ExcelWriter(nome_arquivo, engine='openpyxl',
                        mode='a' if arquivo_existe else 'w',
                        if_sheet_exists='overlay') as writer:

        # Pessoa Física
        if lista_pf:
            df_pf = pd.DataFrame(lista_pf)

            # Tenta pegar a última linha da planilha, se existir
            try:
                book = load_workbook(nome_arquivo)
                if 'Pessoa Física' in book.sheetnames:
                    start_row_pf = book['Pessoa Física'].max_row
                else:
                    start_row_pf = 0
            except:
                start_row_pf = 0

            df_pf.to_excel(writer, sheet_name='Pessoa Física',
                           index=False, header=(start_row_pf == 0), startrow=start_row_pf)

        # Pessoa Jurídica
        if lista_pj:
            df_pj = pd.DataFrame(lista_pj)

            try:
                book = load_workbook(nome_arquivo)
                if 'Pessoa Juridica' in book.sheetnames:
                    start_row_pj = book['Pessoa Juridica'].max_row
                else:
                    start_row_pj = 0
            except:
                start_row_pj = 0

            df_pj.to_excel(writer, sheet_name='Pessoa Juridica',
                           index=False, header=(start_row_pj == 0), startrow=start_row_pj)

    print(f'📁 Arquivo "{nome_arquivo}" atualizado com sucesso!')
