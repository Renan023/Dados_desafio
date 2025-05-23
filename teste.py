import pandas as pd 
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

lista = []

def ajuste(planilha):
    for col_cells in planilha.iter_cols():
        max_length = 0 
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        planilha.column_dimensions[col_letter].width = max_length + 2
     

def excel(lista, arquivo='teste.xlsx'):
    df_teste = pd.DataFrame(lista, columns=['Nome'])
    
    if os.path.exists(arquivo):
        workbook = load_workbook(arquivo)
        sheet = workbook['teste00']
        startrow = sheet.max_row
        

    
    # Reabre a planilha para ajustar colunas
    workbook = load_workbook(arquivo)
    sheet = workbook['teste00']
    ajuste(sheet)
    workbook.save(arquivo)    

# Loop principal
while True:
    print('Digite 1 para continuar ou 2 para sair')
    n = int(input('Qual opção? '))
    
    if n == 1:
        digite = input('Nome: ')
        lista.append([digite])
    else:
        print('Saindo')
        break

excel(lista)