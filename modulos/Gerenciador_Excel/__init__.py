from openpyxl import load_workbook
from modulos.Excel_criar import Excel_vazio
from modulos.Excel_cadastro import cadastro_excel
from modulos.Excel_formatacao import ajuste, criar_tabela, alinhar, cabecalho, bordas, travar_tabela
from modulos.Excel_remove import remove

def gerenciar(nome_arquivo, lista_dados, sheet,background, font, nome_tabela):
    
    #cadastra
    cadastro_excel(nome_arquivo, lista_dados, sheet)
    
    wb = load_workbook(nome_arquivo)
    
    if sheet in wb.sheetnames:
        aba = wb[sheet]
        
        ajuste(aba)
        alinhar(aba)
        cabecalho(aba, background,font)
        bordas(aba)
        criar_tabela(aba,nome_tabela)
        travar_tabela(aba)    

    wb.save(nome_arquivo)
    
    remove(nome_arquivo) 

    print(f"{nome_arquivo} atualizado com sucesso")
    
def gerenciador_completo(nome_arquivo,lista_dados,sheet,background, font, nome_tabela):
    
    #chama para criação do arquivo
    Excel_vazio(nome_arquivo)
    #chama o gerenciar
    gerenciar(nome_arquivo,lista_dados,sheet,background,font,nome_tabela)
    
def gerenciador_semcriar(nome_arquivo,lista_dados,sheet,background,font,nome_tabela):
    
    gerenciar(nome_arquivo,lista_dados,sheet,background,font,nome_tabela)