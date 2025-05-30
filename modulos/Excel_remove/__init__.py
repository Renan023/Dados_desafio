from openpyxl import load_workbook

def remove(nome_arquivo):
    
    book = load_workbook(nome_arquivo)
        
    if 'Sheet' in book.sheetnames:
        book.remove(book['Sheet'])
        book.save(nome_arquivo)
        
        print('Aba Sheet removida')
        
    else:
        
        print('Aba Sheet não encontrada')