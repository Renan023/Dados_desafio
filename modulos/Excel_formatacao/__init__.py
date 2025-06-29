from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table,TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl import load_workbook 

def ajuste(sheet):
    #ajuste 
    for col in sheet.iter_cols(min_row=1 ,max_row=sheet.max_row , min_col=1 ,max_col=sheet.max_column):
        max_length = 0 
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length= max(max_length,len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length +6
        
def criar_tabela(sheet,nome_tabela):
    
    if nome_tabela in sheet.tables:
        del sheet.tables[nome_tabela]
    
    total_linhas = sheet.max_row
    total_colunas = sheet.max_column
    
    ultima_linha = get_column_letter(total_colunas)
    intervalo = f"A1:{ultima_linha}{total_linhas}"
    
    tabela = Table(displayName=nome_tabela, ref=intervalo)
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    ) 
    tabela.tableStyleInfo = estilo
    
    sheet.add_table(tabela)
    
    print(f"Tabela {nome_tabela} criado com sucesso")
    
def alinhar(sheet):
        
        alinhamento = Alignment(horizontal="center",vertical="center")
        
        for row in sheet.iter_rows(min_row = 1, max_row= sheet.max_row, min_col = 1, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = alinhamento
                
def cabecalho(sheet, background,font):

    background = PatternFill(fill_type="solid", fgColor=background)
    texto = Font(color=font)

    for cell in sheet[1]:
        cell.fill = background
        cell.font = texto
        
def bordas(sheet):
    
    borda = Border (
        left=Side(style='thin', color='1F4E78'),
        right=Side(style='thin',color='1F4E78'),
        top=Side(style='thin',color='1F4E78'),
        bottom=Side(style='thin',color='1F4E78')
    )
    
    for row in sheet.iter_rows(min_row = 1, max_row = sheet.max_row, min_col = 1, max_col = sheet.max_column):
        for cell in row:
            cell.border = borda
            
def congelar_cabecalho(sheet):
    #congelar o cabeçalho
    sheet.freeze_panes = 'A2'
    
def travar_tabela(sheet):
    
    try:
        for row in sheet.iter_rows():
            for cell in row:
                cell.protection = cell.protection.copy(locked= True)
                
        sheet.protection.sheet = True
        sheet.protection.set_password('1234')
        
        print(f"Planilha {sheet} protegida com sucesso")
        
    except Exception as e:
        
        print(f"Erro ao proteger {sheet} : {e}")
        
def Copiar(nome_arquivo, sheet, aba_destino):
    
    wb = load_workbook(nome_arquivo)
    
    try:
        
        if aba_destino not in wb.sheetnames:
            destino = wb.create_sheet(aba_destino)
            
        else:
            
            destino = wb[aba_destino]
        
        origem = wb[sheet]
        
        for linha in origem.iter_rows():
            for celula in linha:
                destino[celula.coordinate].value = celula.value
                
        wb.save(nome_arquivo)
        
        print(f"Aba {sheet} copiada com sucesso ")
        
    except Exception as e:
        
        print(f"Não foi possível copiar {sheet} erro: {e}")
        
        