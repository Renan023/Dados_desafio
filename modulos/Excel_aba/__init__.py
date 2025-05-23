from openpyxl import load_workbook
from modulos.Excel_formatacao import ajuste

def create_aba(nome_arquivo,sheet):
    
    #vai reabrir o arquivo e fazer o ajuste de linha     
    wb = load_workbook(nome_arquivo)
        
    if sheet in wb.sheetnames:
        ajuste(wb[sheet])
        
    wb.save(nome_arquivo)
    print(f"{nome_arquivo} atualizado com sucesso")