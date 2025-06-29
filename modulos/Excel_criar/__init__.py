import os 
from openpyxl import Workbook 
from openpyxl import load_workbook
def Excel_vazio(nome_arquivo):
    
    if not os.path.exists(nome_arquivo):
        wb = Workbook()

        wb.save(nome_arquivo)
        
        print(f"{nome_arquivo} criado com sucesso")
    else:
        print(f"{nome_arquivo} já existente")

def planilha_nova(nome_arquivo, aba_destino):
    
    wb = load_workbook(nome_arquivo)
    
    try:
            
        criar = wb[aba_destino] if aba_destino in wb.sheetnames else wb.create_sheet(aba_destino)

        print(f"Planilha {aba_destino} com sucesso ")
        
    except Exception as e:
        
        print(f"Erro ao criar planilha {aba_destino} : {e}")
        
    wb.save(nome_arquivo)