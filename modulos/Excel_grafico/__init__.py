from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

def GraficoB(nome_arquivo, origem, aba_grafico):
    wb = load_workbook(nome_arquivo)

    aba_dados = wb[origem]

    if aba_grafico not in wb.sheetnames:
        aba_destino = wb.create_sheet(aba_grafico)
    else:
        aba_destino = wb[aba_grafico]

    # Referência para os dados e categorias
    dados = Reference(aba_destino, min_col=2, max_col=aba_destino.max_column, min_row=1, max_row=aba_destino.max_row)
    categorias = Reference(aba_destino, min_col=1, min_row=2, max_row=aba_destino.max_row)

    # Criando gráfico
    grafico = BarChart()
    grafico.title = "Gráfico de Análise - Selecione os dados"
    grafico.add_data(dados, titles_from_data=True)
    grafico.set_categories(categorias)

    # Adicionando o gráfico na aba correta
    aba_destino.add_chart(grafico, "E5")

    wb.save(nome_arquivo)
