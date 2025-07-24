from openpyxl.styles import Font
from openpyxl import load_workbook
from modulos.Excel_criar import Planilha_nova
from modulos.Excel_formatacao import ajuste, alinhar, RemoverGridlines
from modulos.Grafico import GraficoDashboard
import pandas as pd 

def DashBoard(nome_arquivo, aba_dash,Graficotitle, destino,start_row):
    
    wb = load_workbook(nome_arquivo)
    
    if destino not in wb.sheetnames:
        print(f"\033[31mAba {destino} não encontrada\033[m")
        return
    
    aba_dados = wb[destino]
    
    if aba_dash not in wb.sheetnames:
        Planilha_nova(nome_arquivo,aba_dash)
        wb = load_workbook(nome_arquivo)
    
    aba_dashboard = wb[aba_dash]
        
    data = list(aba_dados.values)
    colunas = data[0]
    dados = data[1:]
    df = pd.DataFrame(dados, columns = colunas)
    
    aba_dashboard["I1"] = "DASHBOARD DE CLIENTES PESSOA FÍSICA"
    aba_dashboard["I1"].font = Font(bold=True, size=16, color="1F4E79")
    
    total = len(df)
    aba_dashboard["A5"] = "Total de clientes"
    aba_dashboard["B5"] = total
    aba_dashboard["A5"].font = Font(bold=True)
    aba_dashboard["B5"].font = Font(bold=True)
    
    if "Data" in df.columns:
        df["Data"] = pd.to_datetime(df["Data"])
        clientes_mensais = df.groupby(df["Data"].dt.to_period("M")).size()
        
        row = 7
        
        aba_dashboard[f"A{row}"] = "Clientes por Mês"
        aba_dashboard[f"A{row}"].font = Font(bold=True)
        row += 1 
        
        for periodo, qtd in clientes_mensais.items():
            aba_dashboard[f"A{row}"] = str(periodo)
            aba_dashboard[f"B{row}"] = qtd
            aba_dashboard[f"B{row}"].font = Font(color="173EC1")
            row += 1
            
    alinhar(aba_dashboard)
    ajuste(aba_dashboard)
    RemoverGridlines(aba_dashboard)
    GraficoDashboard(aba_dashboard,Graficotitle,start_row=start_row)
    
    wb.save(nome_arquivo)
    
    print("Dashboard criado/preenchido com sucesso!")
    

