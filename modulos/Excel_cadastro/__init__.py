import pandas as pd 
from openpyxl import load_workbook

def cadastro_excel(nome_arquivo,lista_dados,sheet):
    
    while True:
        try:   
            with pd.ExcelWriter(nome_arquivo, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer :
                
                if lista_dados:
                    df_pf = pd.DataFrame(lista_dados)
                    book = load_workbook(nome_arquivo)
                    if sheet in book.sheetnames:
                        start_row_pf = book[sheet].max_row
                    else:
                        start_row_pf = 0
                    df_pf.to_excel(writer, sheet_name=sheet,index=False,header=(start_row_pf==0),startrow=start_row_pf)
            break
        
        except PermissionError:
            
            print(f"\033[31mArquivo {nome_arquivo} aberto ou corrompido\033[m")
            break